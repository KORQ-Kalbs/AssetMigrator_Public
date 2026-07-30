"""
report_generator.py

Writes MigrationReport_<timestamp>.xlsx summarising what happened to every
asset code processed in the run: uploaded / skipped / conflict / error,
the resulting Drive URL, which sheet rows were updated, and timing.
"""

from datetime import datetime
import os

import openpyxl
from openpyxl.styles import Font, PatternFill

import config

STATUS_COLORS = {
    "SUCCESS": "C6EFCE",
    "SKIPPED": "FFEB9C",
    "CONFLICT": "FFC7CE",
    "ERROR": "FFC7CE",
}


def write_report(results: list[dict]) -> str:
    """
    results: list of dicts with keys:
        asset_code, asset_name, status, rows_updated, url, message, seconds
    """
    os.makedirs(config.REPORT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(config.REPORT_DIR, f"MigrationReport_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Migration Report"

    headers = ["Asset Code", "Asset Name", "Status", "Rows Updated", "Drive URL", "Message", "Seconds"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in results:
        ws.append([
            r.get("asset_code", ""),
            r.get("asset_name", ""),
            r.get("status", ""),
            ", ".join(str(x) for x in r.get("rows_updated", [])),
            r.get("url", ""),
            r.get("message", ""),
            round(r.get("seconds", 0), 2),
        ])
        color = STATUS_COLORS.get(r.get("status", ""))
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=ws.max_row, column=3).fill = fill

    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

    # Summary sheet
    summary = wb.create_sheet("Summary")
    counts = {}
    for r in results:
        counts[r.get("status", "")] = counts.get(r.get("status", ""), 0) + 1
    summary.append(["Status", "Count"])
    for status, count in counts.items():
        summary.append([status, count])
    summary.append(["TOTAL", len(results)])

    wb.save(path)
    return path
