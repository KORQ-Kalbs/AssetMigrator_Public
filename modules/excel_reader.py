"""
excel_reader.py

Reads the source .xlsx export and extracts embedded images, matching each
image to its Asset Code / Asset Name via the image's cell ANCHOR position
(row/column), not by list order or sequential guessing.

Why anchor-based matching:
Embedded images in openpyxl (`worksheet._images`) are NOT returned in
row order - they're returned in the order they were inserted into the
XML. Each image, however, carries a `OneCellAnchor` (or TwoCellAnchor)
with an exact 0-indexed row/column. That anchor is the reliable source
of truth for "which row does this image belong to".
"""

from dataclasses import dataclass
from typing import Optional
import xml.etree.ElementTree as ET
import zipfile
import openpyxl
from loguru import logger


@dataclass
class ExtractedAsset:
    row_number: int          # 1-indexed row in the SOURCE excel file
    asset_code: str
    asset_name: Optional[str]
    image_bytes: bytes
    image_ext: str            # e.g. "jpeg", "png"


def _build_header_map(ws, header_row: int) -> dict:
    """Map header text (stripped) -> 1-indexed column number."""
    headers = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column
    return headers


def _extract_images_via_zip(xlsx_path: str) -> list[tuple[int, bytes, str]]:
    """
    Fallback image extractor using zipfile & XML parsing directly from the .xlsx archive.
    Returns list of (1-indexed row number, image_bytes, image_ext).
    Handles relative paths like ../drawings/drawing1.xml that openpyxl fails to resolve.
    """
    extracted = []
    try:
        with zipfile.ZipFile(xlsx_path, "r") as z:
            namelist = z.namelist()
            drawing_files = [f for f in namelist if "xl/drawings/drawing" in f and f.endswith(".xml")]
            for df in drawing_files:
                df_name = df.split("/")[-1]
                rel_path = f"xl/drawings/_rels/{df_name}.rels"
                if rel_path not in namelist:
                    continue

                rels_tree = ET.fromstring(z.read(rel_path))
                rid_map = {}
                for r in rels_tree:
                    target = r.attrib.get("Target", "")
                    target_clean = target.replace("../media/", "xl/media/").lstrip("/")
                    if not target_clean.startswith("xl/"):
                        target_clean = "xl/" + target_clean
                    rid_map[r.attrib["Id"]] = target_clean

                tree = ET.fromstring(z.read(df))
                ns = {
                    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                }
                anchors = tree.findall(".//xdr:oneCellAnchor", ns) + tree.findall(".//xdr:twoCellAnchor", ns)
                for a in anchors:
                    row_elem = a.find("./xdr:from/xdr:row", ns)
                    if row_elem is None or row_elem.text is None:
                        continue
                    row = int(row_elem.text) + 1  # 0-indexed -> 1-indexed
                    blip = a.find(".//a:blip", ns)
                    if blip is None:
                        continue
                    embed = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                    if not embed or embed not in rid_map:
                        continue
                    img_path = rid_map[embed]
                    if img_path in namelist:
                        img_bytes = z.read(img_path)
                        ext = img_path.split(".")[-1].lower()
                        extracted.append((row, img_bytes, ext))
    except Exception as e:
        logger.warning(f"Zipfile direct extraction fallback failed: {e}")
    return extracted


def extract_assets_with_images(
    xlsx_path: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    asset_code_header: str = "Asset Code",
    asset_name_header: str = "Asset Name",
) -> list[ExtractedAsset]:
    """
    Open the source workbook and return one ExtractedAsset per embedded
    image, keyed to the Asset Code / Asset Name found in that image's
    anchor row.
    """
    logger.info(f"Reading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    headers = _build_header_map(ws, header_row)
    if asset_code_header not in headers:
        raise ValueError(
            f"Could not find header '{asset_code_header}' on row {header_row} "
            f"of sheet '{ws.title}'. Found headers: {list(headers.keys())}"
        )
    code_col = headers[asset_code_header]
    name_col = headers.get(asset_name_header)

    images = getattr(ws, "_images", [])
    logger.info(f"Found {len(images)} embedded image(s) via openpyxl in '{ws.title}'")

    raw_items = []
    if images:
        for img in images:
            anchor = img.anchor
            anchor_from = getattr(anchor, "_from", None)
            if anchor_from is None:
                continue
            row = anchor_from.row + 1
            try:
                raw_items.append((row, img._data(), (img.format or "png").lower()))
            except Exception:
                continue
    else:
        logger.info("Using Zipfile XML direct extraction fallback...")
        raw_items = _extract_images_via_zip(xlsx_path)

    results: list[ExtractedAsset] = []
    seen_codes = set()

    for actual_row, image_bytes, image_ext in raw_items:
        asset_code_cell = ws.cell(row=actual_row, column=code_col).value
        asset_code = str(asset_code_cell).strip() if asset_code_cell is not None else None

        asset_name = None
        if name_col:
            name_cell = ws.cell(row=actual_row, column=name_col).value
            asset_name = str(name_cell).strip() if name_cell is not None else None

        if not asset_code:
            logger.error(
                f"Row {actual_row}: image found but no Asset Code in that row - skipping"
            )
            continue

        if asset_code in seen_codes:
            logger.warning(
                f"Asset Code '{asset_code}' has more than one image in row {actual_row} - keeping first"
            )
            continue
        seen_codes.add(asset_code)

        results.append(
            ExtractedAsset(
                row_number=actual_row,
                asset_code=asset_code,
                asset_name=asset_name,
                image_bytes=image_bytes,
                image_ext=image_ext,
            )
        )

    logger.info(f"Successfully extracted {len(results)} asset image(s)")
    return results
